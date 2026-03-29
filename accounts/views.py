import os
import random, string
from datetime import datetime
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.http import JsonResponse
import openpyxl
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mass_mail
from django.conf import settings
from django.http import HttpResponse
from .models import User, Team, TeamMember, DocumentSlot, TeamSubmission
from django.db.models import Count
from django.db import transaction
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from django.db.models import Q
import json
import io
from django.views.decorators.csrf import csrf_exempt

def check_reg_number(request):
    reg = request.GET.get('reg', None)
    data = {
        'is_taken': TeamMember.objects.filter(reg_number__iexact=reg).exists()
    }
    return JsonResponse(data)

def portal_gatekeeper(request):
    """The 4-panel landing page."""
    return render(request, 'accounts/portal_gatekeeper.html')

# accounts/views.py
from django.db.models import Count
import random, string

def register_team(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        l_name = request.POST.get('leader_name')

        # 1. Password Match Check
        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, 'accounts/register.html')

        # 2. Email Existence Check
        if User.objects.filter(email=email).exists():
            messages.error(request, "This email is already registered.")
            return render(request, 'accounts/register.html')

        # 3. Collect and Validate Member Data
        reg_numbers = []
        member_data = []
        for i in range(1, 5):
            name = request.POST.get(f'm{i}_name')
            reg = request.POST.get(f'm{i}_reg')
            if name and reg:
                # Format Check: Ensure it starts with AIK
                if not reg.startswith('AIK'):
                    messages.error(request, f"Invalid format for {reg}. Must start with 'AIK'.")
                    return render(request, 'accounts/register.html')
                reg_numbers.append(reg)
                member_data.append({'name': name, 'reg': reg})

        # 4. Check for Duplicate Register Numbers in Database
        existing_members = TeamMember.objects.filter(reg_number__in=reg_numbers)
        if existing_members.exists():
            duplicate_regs = ", ".join([m.reg_number for m in existing_members])
            messages.error(request, f"Register number(s) {duplicate_regs} already exist in our database.")
            return render(request, 'accounts/register.html')

        try:
            with transaction.atomic():
                # 1. Create User
                user = User.objects.create_user(email=email, username=email, password=password, role='TEAM')
                
                # 2. Allocate Guide (Load Balancing Logic)
                guides = User.objects.filter(role='GUIDE').annotate(
                    t_count=Count('team_profile')
                ).order_by('t_count','id')
                
                assigned_guide = guides.first()
                
                if not assigned_guide:
                    raise ValueError("No Faculty Guides are currently available.")

                # 3. Create Team (save() handles ID generation)
                team = Team.objects.create(
                    user=user,
                    guide=assigned_guide,
                    leader_name=l_name
                )

                # 4. Save Members
                for m in member_data:
                    TeamMember.objects.create(
                        team=team,
                        name=m['name'],
                        reg_number=m['reg'],
                        is_leader=(m['name'] == l_name)
                    )

            # Success!
            messages.success(request, f"Successfully Registered! Your Team ID is: {team.team_id}")
            return redirect('team_login')

        except Exception as e:
            print(f"Error during registration: {e}")
            messages.error(request, f"Error: {str(e)}")
            return render(request, 'accounts/register.html')

    return render(request, 'accounts/register.html')

def coordinator_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user and user.role == 'COORDINATOR':
            login(request, user)
            return redirect('coordinator_dashboard')
        messages.error(request, "Invalid credentials or not a Coordinator.")
    return render(request, 'accounts/login_coordinator.html')

def hod_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user and user.role == 'HOD':
            login(request, user)
            return redirect('hod_dashboard') # Ensure this URL name exists
        messages.error(request, "Invalid credentials or not an HOD.")
    return render(request, 'accounts/login_hod.html')

def guide_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user and user.role == 'GUIDE':
            login(request, user)
            return redirect('guide_dashboard') # Ensure this URL name exists
        messages.error(request, "Invalid credentials or not a Guide.")
    return render(request, 'accounts/login_guide.html')

def team_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        user = authenticate(request, username=email, password=password)
        if user and user.role == 'TEAM':
            login(request, user)
            return redirect('team_dashboard')
        messages.error(request, "Invalid credentials or not a Student Team.")
    return render(request, 'accounts/team_login.html')

from django.db.models import Q

@login_required
def team_dashboard(request):
    # Ensure only logged-in teams can see this
    if not request.user.is_authenticated or request.user.role != 'TEAM':
        return redirect('team_login')

    try:
        # Access the team profile linked to the user
        team = request.user.student_profile
    except Team.DoesNotExist:
        messages.error(request, "Team profile not found. Please contact the Coordinator.")
        return redirect('team_login')

    members = team.members.all()
    
    # Fetch active slots (Deadlines or Reviews)
    active_slots = DocumentSlot.objects.filter(
        Q(is_active=True) | Q(review_date__isnull=False)
    ).distinct().order_by('deadline')

    # --- UPDATED: Create a mapping of slot types to submission objects ---
    submissions = TeamSubmission.objects.filter(team=team).select_related('slot')
    
    # This dictionary allows the 'get_item' filter in the HTML to find 
    # the specific submission (and its is_approved status) for each slot.
    submissions_map = {sub.slot.slot_type: sub for sub in submissions}
    
    # Legacy list for simple 'Submitted' vs 'Pending' checks
    submitted_slots = [sub.slot_id for sub in submissions]

    context = {
        'team': team,
        'members': members,
        'guide': team.guide,
        'active_slots': active_slots,
        'submissions_map': submissions_map,  # New: Required for approval status
        'submitted_slots': submitted_slots,  # Existing: Required for 'Submitted' text
        'now': timezone.now(),
    }
    return render(request, 'accounts/team_dashboard.html', context)

@login_required
def upload_document(request, slot_id):
    if request.method == 'POST' and request.FILES.get('doc_file'):
        try:
            team = request.user.student_profile
            slot = DocumentSlot.objects.get(id=slot_id)
            uploaded_file = request.FILES['doc_file']

            # Check if a submission already exists for this team and slot
            existing_submission = TeamSubmission.objects.filter(team=team, slot=slot).first()

            if existing_submission:
                # Replace the existing document
                if existing_submission.file:
                    existing_submission.file.delete(save=False)
                
                existing_submission.file = uploaded_file
                existing_submission.submitted_at = timezone.now()
                # Optional: existing_submission.is_approved = False
                existing_submission.save()
                messages.success(request, f"Document for {slot.title} has been updated.")
            else:
                # Create a new submission record
                TeamSubmission.objects.create(
                    team=team,
                    slot=slot,
                    file=uploaded_file
                )
                messages.success(request, f"File for {slot.title} uploaded successfully.")
                
        except Exception as e:
            messages.error(request, f"Upload failed: {str(e)}")

    return redirect('team_dashboard')

@login_required
def coordinator_dashboard(request):
    if request.user.role != 'COORDINATOR':
        return redirect('coordinator_login')

    teams = Team.objects.all().prefetch_related('members').order_by('-created_at')
    
    slots = DocumentSlot.objects.all()
    slots_dict = {slot.slot_type: slot for slot in slots}

    from collections import defaultdict
    submissions_raw = TeamSubmission.objects.all().select_related('slot')
    team_submissions_map = defaultdict(dict)
    for sub in submissions_raw:
        team_submissions_map[sub.team_id][sub.slot.slot_type] = sub.file.url 

    if 'export_excel' in request.GET:
        # ... (Your existing Excel export code remains the same)
        pass

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'approve_title':
            team = get_object_or_404(Team, pk=request.POST.get('team_id'))
            team.project_title = request.POST.get('project_title')
            team.is_approved = True 
            team.save()
            messages.success(request, f"Project title approved for Team {team.team_id}.")
            return redirect('coordinator_dashboard')

        elif action == 'edit_title':
            team = get_object_or_404(Team, pk=request.POST.get('team_id'))
            team.is_approved = False 
            team.save()
            return redirect('coordinator_dashboard')

        elif action == 'toggle_approval':
            team_id = request.POST.get('team_id')
            slot_type = request.POST.get('slot_type')
            
            submission = TeamSubmission.objects.filter(
                team__team_id=team_id, 
                slot__slot_type=slot_type
            ).last()

            if submission:
                submission.is_approved = not submission.is_approved
                submission.save()
                status = "Approved" if submission.is_approved else "Unapproved"
                messages.success(request, f"{slot_type} set to {status} for Team {team_id}.")
                
            return redirect('coordinator_dashboard')

        elif action == 'set_deadline':
            slot_type = request.POST.get('slot_type')
            notification_msg = ""
            
            if slot_type in ['REV1', 'REV2', 'PPT1', 'PPT2']:
                date_val = request.POST.get('review_date') or request.POST.get('specific_review_date')
                DocumentSlot.objects.update_or_create(
                    slot_type=slot_type,
                    defaults={'title': f"{slot_type} Presentation", 'review_date': date_val, 'is_active': True, 'deadline': None}
                )
                notification_msg = f"{slot_type} date set."
            else:
                date_val = request.POST.get('deadline_date')
                titles = {'PROPOSAL': 'Project Proposal', 'ABSTRACT': 'Abstract', 'SRS': 'SRS Document','FINAL_PPT': 'Final PPT', 'REPORT': 'Final Report'}
                title = titles.get(slot_type, "Document Submission")
                DocumentSlot.objects.update_or_create(
                    slot_type=slot_type,
                    defaults={'title': title, 'deadline': date_val, 'is_active': True, 'review_date': None}
                )
                notification_msg = "Deadline updated."

            messages.success(request, f"Schedule for {slot_type} has been successfully updated!")
            return redirect('coordinator_dashboard')
        
        elif action == 'delete_deadline':
            slot_type = request.POST.get('slot_type')
            slot = DocumentSlot.objects.filter(slot_type=slot_type).first()
            if slot:
                slot.delete()
                messages.success(request, "Deadline removed.")
            return redirect('coordinator_dashboard')

    context = {
        'teams': teams,
        'slots_dict': slots_dict,
        'team_submissions_map': dict(team_submissions_map),
        'now': timezone.now(),
        'members': TeamMember.objects.all().order_by('reg_number'),
    }
    return render(request, 'accounts/coordinator_dashboard.html', context)

@login_required
def evaluate_r1_batch_coordinator(request):
    members = TeamMember.objects.all().order_by('team__created_at','reg_number')
    if request.method == 'POST':
        for m in members:
            m.r1_c_comp = float(request.POST.get(f'comp_{m.id}', 0))
            m.r1_c_func = float(request.POST.get(f'func_{m.id}', 0))
            m.r1_c_pres = float(request.POST.get(f'pres_{m.id}', 0))
            m.r1_c_oral = float(request.POST.get(f'oral_{m.id}', 0))
            m.r1_c_know = float(request.POST.get(f'know_{m.id}', 0))
            m.r1_c_absent = f'absent_{m.id}' in request.POST
            m.r1_coord_total = m.r1_c_comp + m.r1_c_func + m.r1_c_pres + m.r1_c_oral + m.r1_c_know
            m.save()
        messages.success(request, "Review 1 Evaluation sheet updated successfully.")
        return redirect('coordinator_dashboard')
    return render(request, 'accounts/batch_sheet_r1_coordinator.html', {'members': members})

@login_required
def evaluate_r2_batch_coordinator(request):
    members = TeamMember.objects.all().order_by('team__created_at','reg_number')
    if request.method == 'POST':
        for m in members:
            m.r2_c_comp = float(request.POST.get(f'comp_{m.id}', 0))
            m.r2_c_func = float(request.POST.get(f'func_{m.id}', 0))
            m.r2_c_pres = float(request.POST.get(f'pres_{m.id}', 0))
            m.r2_c_oral = float(request.POST.get(f'oral_{m.id}', 0))
            m.r2_c_know = float(request.POST.get(f'know_{m.id}', 0))
            m.r2_c_absent = f'absent_{m.id}' in request.POST
            m.r2_coord_total = m.r2_c_comp + m.r2_c_func + m.r2_c_pres + m.r2_c_oral + m.r2_c_know
            m.save()
        messages.success(request, "Review 2 Evaluation sheet updated successfully.")
        return redirect('coordinator_dashboard')
    return render(request, 'accounts/batch_sheet_r2_coordinator.html', {'members': members})

@login_required
def evaluate_s2_batch_coordinator(request):
    members = TeamMember.objects.all().order_by('team__created_at','reg_number')
    if request.method == 'POST':
        for m in members:
            m.s2_teamwork = float(request.POST.get(f'teamwork_{m.id}', 0))
            m.s2_tech_know = float(request.POST.get(f'tech_{m.id}', 0))
            m.s2_regularity = float(request.POST.get(f'reg_{m.id}', 0))
            m.s2_total = m.s2_teamwork + m.s2_tech_know + m.s2_regularity
            m.save()
        messages.success(request, "Institutional Sheet 2 updated.")
        return redirect('coordinator_dashboard')
    return render(request, 'accounts/batch_sheet_s2_coordinator.html', {'members': members})

@login_required
def evaluate_report_batch_coordinator(request):
    members = TeamMember.objects.all().order_by('team__created_at','reg_number')
    if request.method == 'POST':
        for m in members:
            m.report_coord = float(request.POST.get(f'report_{m.id}', 0))
            m.save()
        messages.success(request, "Report marks have been updated successfully.")
        return redirect('coordinator_dashboard')
    return render(request, 'accounts/batch_sheet_report_coordinator.html', {'members': members})

@login_required
def evaluate_attendance_batch_coordinator(request):
    members = TeamMember.objects.all().order_by('team__created_at','reg_number')
    if request.method == 'POST':
        for m in members:
            m.attendance_marks = float(request.POST.get(f'attendance_{m.id}', 0))
            m.save()
        messages.success(request, "Student attendance marks have been updated successfully.")
        return redirect('coordinator_dashboard')
    return render(request, 'accounts/batch_sheet_attendance_coordinator.html', {'members': members})

from django.contrib import messages

# HOD Review 1 View
@login_required
def evaluate_r1_hod(request):
    members = TeamMember.objects.all().select_related('team').order_by('team__created_at', 'reg_number')
    if request.method == 'POST':
        for m in members:
            m.r1_h_comp = float(request.POST.get(f'comp_{m.id}', 0))
            m.r1_h_func = float(request.POST.get(f'func_{m.id}', 0))
            m.r1_h_pres = float(request.POST.get(f'pres_{m.id}', 0))
            m.r1_h_oral = float(request.POST.get(f'oral_{m.id}', 0))
            m.r1_h_know = float(request.POST.get(f'know_{m.id}', 0))
            m.r1_h_absent = f'absent_{m.id}' in request.POST
            m.r1_hod_total = m.r1_h_comp + m.r1_h_func + m.r1_h_pres + m.r1_h_oral + m.r1_h_know
            m.save()
        messages.success(request, "HOD Review 1 marks updated.")
        return redirect('hod_dashboard')
    return render(request, 'accounts/evaluate_r1_hod.html', {'members': members})

# HOD Review 2 View
@login_required
def evaluate_r2_hod(request):
    members = TeamMember.objects.all().select_related('team').order_by('team__created_at', 'reg_number')
    if request.method == 'POST':
        for m in members:
            m.r2_h_comp = float(request.POST.get(f'comp_{m.id}', 0))
            m.r2_h_func = float(request.POST.get(f'func_{m.id}', 0))
            m.r2_h_pres = float(request.POST.get(f'pres_{m.id}', 0))
            m.r2_h_oral = float(request.POST.get(f'oral_{m.id}', 0))
            m.r2_h_know = float(request.POST.get(f'know_{m.id}', 0))
            m.r2_h_absent = f'absent_{m.id}' in request.POST
            m.r2_hod_total = m.r2_h_comp + m.r2_h_func + m.r2_h_pres + m.r2_h_oral + m.r2_h_know
            m.save()
        messages.success(request, "HOD Review 2 marks updated.")
        return redirect('hod_dashboard')
    return render(request, 'accounts/evaluate_r2_hod.html', {'members': members})

# HOD Report Mark Sheet View
@login_required
def evaluate_report_hod(request):
    members = TeamMember.objects.all().select_related('team').order_by('team__created_at', 'reg_number')
    if request.method == 'POST':
        for m in members:
            m.report_hod = float(request.POST.get(f'report_{m.id}', 0))
            m.save()
        messages.success(request, "HOD Report marks updated.")
        return redirect('hod_dashboard')
    return render(request, 'accounts/evaluate_report_hod.html', {'members': members})

@login_required
def evaluate_r1_guide(request):
    # SECURITY: Filter members so the guide only sees their assigned teams
    members = TeamMember.objects.filter(
        team__guide=request.user
    ).select_related('team').order_by('team__created_at', 'reg_number')
    
    if request.method == 'POST':
        for m in members:
            # Map inputs to GUIDE fields (r1_g_comp, r1_g_func, etc.)
            m.r1_g_comp = float(request.POST.get(f'comp_{m.id}', 0) or 0)
            m.r1_g_func = float(request.POST.get(f'func_{m.id}', 0) or 0)
            m.r1_g_pres = float(request.POST.get(f'pres_{m.id}', 0) or 0)
            m.r1_g_oral = float(request.POST.get(f'oral_{m.id}', 0) or 0)
            m.r1_g_know = float(request.POST.get(f'know_{m.id}', 0) or 0)
            m.r1_g_absent = f'absent_{m.id}' in request.POST
            
            # Update the persistent GUIDE total field
            m.r1_guide_total = m.r1_g_comp + m.r1_g_func + m.r1_g_pres + m.r1_g_oral + m.r1_g_know
            m.save()
            
        messages.success(request, "Guide Review 1 marks updated successfully.")
        return redirect('guide_dashboard')
        
    return render(request, 'accounts/evaluate_r1_guide.html', {'members': members})

@login_required
def evaluate_r2_guide(request):
    # Security: Filter so the guide only sees their assigned teams
    members = TeamMember.objects.filter(
        team__guide=request.user
    ).select_related('team').order_by('team__created_at', 'reg_number')
    
    if request.method == 'POST':
        for m in members:
            # Map inputs to GUIDE R2 fields
            m.r2_g_comp = float(request.POST.get(f'comp_{m.id}', 0) or 0)
            m.r2_g_func = float(request.POST.get(f'func_{m.id}', 0) or 0)
            m.r2_g_pres = float(request.POST.get(f'pres_{m.id}', 0) or 0)
            m.r2_g_oral = float(request.POST.get(f'oral_{m.id}', 0) or 0)
            m.r2_g_know = float(request.POST.get(f'know_{m.id}', 0) or 0)
            m.r2_g_absent = f'absent_{m.id}' in request.POST
            
            # Update the persistent GUIDE R2 total field
            m.r2_guide_total = m.r2_g_comp + m.r2_g_func + m.r2_g_pres + m.r2_g_oral + m.r2_g_know
            m.save()
            
        messages.success(request, "Guide Review 2 marks updated successfully.")
        return redirect('guide_dashboard')
        
    return render(request, 'accounts/evaluate_r2_guide.html', {'members': members})

# GUIDE REPORT EVALUATION VIEW
@login_required
def evaluate_report_guide(request):
    # Filter for assigned teams
    members = TeamMember.objects.filter(
        team__guide=request.user
    ).select_related('team').order_by('team__created_at', 'reg_number')
    
    if request.method == 'POST':
        for m in members:
            # Map input to report_guide field
            val = request.POST.get(f'report_{m.id}', 0)
            m.report_guide = float(val) if val else 0.0
            m.save()
            
        messages.success(request, "Guide Report marks updated successfully.")
        return redirect('guide_dashboard')
        
    return render(request, 'accounts/evaluate_report_guide.html', {'members': members})

@login_required
def update_review_date(request):
    if request.method == 'POST' and request.user.role == 'COORDINATOR':
        s_type = request.POST.get('slot_type')
        date_str = request.POST.get('specific_review_date')
        
        if s_type and date_str:
            # Parse the specific date selected (e.g., March 3)
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            # Validation: No past dates
            if selected_date < timezone.now().date():
                messages.error(request, "Cannot schedule a review for a past date.")
                return redirect('coordinator_dashboard')

            # The Full-Time Fix: update or create the record
            slot, created = DocumentSlot.objects.update_or_create(
                slot_type=s_type,
                defaults={
                    'title': 'Review 1 Date' if s_type == 'REV1' else 'Review 2 Date',
                    'review_date': selected_date,
                    'is_active': True,
                    'deadline': None # Now allowed because of the model change
                }
            )
            messages.success(request, f"Review scheduled for {selected_date}")
            
    return redirect('coordinator_dashboard')

# --- REUSABLE EXCEL HELPER ---

def export_to_excel(queryset, filename, sheet_name, headers, row_callback):
    """Generates an Excel response using openpyxl."""
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = str(sheet_name)[:31]
    
    # Add Bold Headers
    ws.append(headers)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    
    for obj in queryset:
        row_data = row_callback(obj)
        # Convert any None, datetime, or complex objects to strings
        sanitized_row = [str(item) if item is not None else "" for item in row_data]
        ws.append(sanitized_row)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    ws.freeze_panes = 'A2'    
    wb.save(response)
    return response

# --- CONSOLIDATED PDF & EXCEL REPORTS ---

def generate_branded_pdf(template_src, context_dict, filename):
    logo_path = os.path.join(settings.BASE_DIR, 'static', 'assets', 'branding', 'logo.png')
    context_dict['logo_path'] = logo_path
    html = render_to_string(template_src, context_dict)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

@login_required
def report_r1_consolidated(request):
    members = TeamMember.objects.all().order_by('reg_number')
    if request.GET.get('format') == 'excel':
        return export_to_excel(members, "R1_Consolidated", "Review 1", 
                              ['Reg No', 'Name', 'Guide', 'HOD', 'Coord', 'Total (40)'],
                              lambda m: [m.reg_number, m.name, m.r1_guide_total, m.r1_hod_total, m.r1_coord_total, m.r1_consolidated_40])
    return generate_branded_pdf('accounts/pdf_r1_cons.html', {'members': members, 'title': 'REVIEW 1 CONSOLIDATED'}, 'R1_Consolidated')

@login_required
def report_r2_consolidated(request):
    members = TeamMember.objects.all().order_by('reg_number')
    if request.GET.get('format') == 'excel':
        return export_to_excel(members, "R2_Consolidated", "Review 2", 
                              ['Reg No', 'Name', 'Guide', 'HOD', 'Coord', 'Total (40)'],
                              lambda m: [m.reg_number, m.name, m.r2_guide_total, m.r2_hod_total, m.r2_coord_total, m.r2_consolidated_40])
    return generate_branded_pdf('accounts/pdf_r2_cons.html', {'members': members, 'title': 'REVIEW 2 CONSOLIDATED'}, 'R2_Consolidated')

@login_required
def report_avg_evaluation_consolidated(request):
    members = TeamMember.objects.all().order_by('reg_number')
    if request.GET.get('format') == 'excel':
        return export_to_excel(members, "Avg_Evaluation", "Average Eval", 
                              ['Reg No', 'Name', 'R1 Cons', 'R2 Cons', 'Avg (40)'],
                              lambda m: [m.reg_number, m.name, m.r1_consolidated_40, m.r2_consolidated_40, m.avg_evaluation_40])
    return generate_branded_pdf('accounts/pdf_avg_eval_cons.html', {'members': members, 'title': 'AVERAGE EVALUATION (40)'}, 'Avg_Evaluation')

@login_required
def report_report_marks_consolidated(request):
    members = TeamMember.objects.all().order_by('reg_number')
    if request.GET.get('format') == 'excel':
        return export_to_excel(members, "Report_Marks", "Report Marks", 
                              ['Reg No', 'Name', 'Guide', 'HOD', 'Coord', 'Consolidated (10)'],
                              lambda m: [m.reg_number, m.name, m.report_guide, m.report_hod, m.report_coord, m.consolidated_report_marks])
    return generate_branded_pdf('accounts/pdf_report_cons.html', {'members': members, 'title': 'REPORT MARKS CONSOLIDATED'}, 'Report_Marks')

@login_required
def report_final_internal(request):
    if request.user.role not in ['COORDINATOR', 'HOD']:
        return redirect('portal_gatekeeper')
    members = TeamMember.objects.all().order_by('reg_number')
    if request.GET.get('format') == 'excel':
        return export_to_excel(members, "Final_Internal_75", "Final Internal", 
                              ['Reg No', 'Name', 'Avg Eval', 'Sheet 2', 'Report', 'Attend', 'Final (75)'],
                              lambda m: [m.reg_number, m.name, m.avg_evaluation_40, m.s2_total, m.consolidated_report_marks, m.attendance_marks, m.final_internal_75])
    return generate_branded_pdf('accounts/pdf_final_internal.html', {'members': members, 'title': 'FINAL INTERNAL MARKS (75)'}, 'Final_Internal_75')

@login_required
def report_team_master_pdf(request):
    if request.user.role not in ['COORDINATOR', 'HOD']:
        return redirect('portal_gatekeeper')
    teams = Team.objects.all().prefetch_related('members').select_related('guide').order_by('team_id')
    
    if request.GET.get('format') == 'excel':
        # Custom logic for Team Master because it uses 'teams' instead of 'members'
        return export_to_excel(teams, "Team_Master_Record", "Teams", 
                              ['Team ID', 'Project Title', 'Guide', 'Members Details'],
                              lambda t: [t.team_id, t.project_title or "Not Set", str(t.guide), ", ".join([f"{m.name} ({m.reg_number})" for m in t.members.all()])])
        
    return generate_branded_pdf('accounts/pdf_master_sheet.html', {'teams': teams, 'title': 'OFFICIAL TEAM MASTER RECORD'}, 'Team_Master_Sheet')

def logout_view(request):
    """Clears the session and redirects to the landing page."""
    from django.contrib.auth import logout
    logout(request)
    return redirect('portal_gatekeeper')

@login_required
def hod_dashboard(request):
    if request.user.role != 'HOD':
        return redirect('hod_login')
    
    # FIX: Fetch slots that have a review_date set
    scheduled_reviews = DocumentSlot.objects.filter(review_date__isnull=False).order_by('review_date')
    
    teams = Team.objects.all().prefetch_related('members').order_by('team_id')
    
    return render(request, 'accounts/hod_dashboard.html', {
        'teams': teams,
        'scheduled_reviews': scheduled_reviews # Ensure this is passed to the template
    })


@login_required
def guide_dashboard(request):
    if request.user.role != 'GUIDE':
        return redirect('guide_login')
    
    scheduled_reviews = DocumentSlot.objects.filter(review_date__isnull=False).order_by('review_date')
    assigned_teams = Team.objects.filter(guide=request.user).prefetch_related('members')
    
    return render(request, 'accounts/guide_dashboard.html', {'teams': assigned_teams,'scheduled_reviews': scheduled_reviews})

@login_required
def delete_review_date(request):
    if request.method == 'POST' and request.user.role == 'COORDINATOR':
        s_type = request.POST.get('slot_type')
        
        # Find the slot and reset only the review_date
        slot = DocumentSlot.objects.filter(slot_type=s_type).first()
        if slot:
            slot.review_date = None
            slot.save()
            messages.warning(request, f"Scheduled date for {slot.title} has been cleared.")
        else:
            messages.error(request, "Review slot not found.")
            
    return redirect('coordinator_dashboard')

def update_project_title(request, team_id):
    if request.method == "POST":
        # Retrieve the specific team or return 404
        team = get_object_or_404(Team, team_id=team_id)
        new_title = request.POST.get('new_title')
        
        if new_title:
            team.project_title = new_title
            # Ensure the team remains marked as approved
            team.is_approved = True 
            team.save()
            messages.success(request, f"Project title for {team.team_id} updated successfully!")
        else:
            messages.error(request, "Title cannot be empty.")
            
    return redirect('coordinator_dashboard')

@login_required
def export_live_evaluation(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            format_type = request.GET.get('format')
            title = body.get('title', 'Evaluation_Draft')
            eval_data = body.get('data', [])
            eval_type = body.get('type', 'standard') 

            # --- HANDLE EXCEL EXPORT ---
            if format_type == 'excel':
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{title}.xlsx"'
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Live Draft"

                # A. Attendance & Report Layouts (3-4 columns)
                if eval_type in ['attendance_coordinator', 'report_coordinator', 'report_guide', 'report_hod']:
                    headers = ['Sl. No', 'Register Number', 'Name', 'Marks (10)']
                    ws.append(headers)
                    for row in eval_data:
                        ws.append([row.get('sl_no'), row.get('reg_no'), row.get('name'), row.get('marks')])

                # B. Institutional S2 Layout
                elif eval_type == 'institutional_s2_coordinator':
                    headers = ['Register Number', 'Name', 'Teamwork (5)', 'Tech Know (5)', 'Regularity (5)', 'Total (15)', 'Absent']
                    ws.append(headers)
                    for row in eval_data:
                        ws.append([row.get('reg_no'), row.get('name'), row.get('teamwork'), row.get('tech_know'), row.get('regularity'), row.get('total'), row.get('absent')])

                # C. All Review Layouts (Coordinator, Guide, HOD)
                else:
                    headers = ['Register Number', 'Name', 'Completion (10)', 'Functionality (5)', 'Presentation (5)', 'Oral (10)', 'Knowledge (10)', 'Total (40)', 'Absent']
                    ws.append(headers)
                    for row in eval_data:
                        ws.append([row.get('reg_no'), row.get('name'), row.get('completion'), row.get('functionality'), row.get('presentation'), row.get('oral'), row.get('knowledge'), row.get('total'), row.get('absent')])

                # Style & Auto-Fit
                for cell in ws[1]: cell.font = openpyxl.styles.Font(bold=True)
                for col in ws.columns:
                    max_len = max([len(str(cell.value)) for cell in col])
                    ws.column_dimensions[col[0].column_letter].width = max_len + 4
                
                wb.save(response)
                return response

            # --- HANDLE PDF EXPORT ---
            elif format_type == 'pdf':
                context = {'title': title.upper(), 'members': eval_data, 'today': timezone.now()}
                
                # 1:1 Mapping to Template Files
                template_map = {
                    'review1_coordinator': 'accounts/pdf_review1_coordinator.html',
                    'review2_coordinator': 'accounts/pdf_review2_coordinator.html',
                    'attendance_coordinator': 'accounts/pdf_attendance_coordinator.html',
                    'report_coordinator': 'accounts/pdf_report_coordinator.html',
                    'institutional_s2_coordinator': 'accounts/pdf_institutional_s2_coordinator.html',
                    'review1_guide': 'accounts/pdf_review1_guide.html',
                    'review2_guide': 'accounts/pdf_review2_guide.html',
                    'report_guide': 'accounts/pdf_report_guide.html',
                    'review1_hod': 'accounts/pdf_review1_hod.html',
                    'review2_hod': 'accounts/pdf_review2_hod.html',
                    'report_hod': 'accounts/pdf_report_hod.html',
                }
                
                template = template_map.get(eval_type, 'accounts/pdf_template.html')
                return generate_branded_pdf(template, context, title)

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return HttpResponse("Invalid Request", status=400)